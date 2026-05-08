import pandas as pd
from openpyxl import load_workbook

wb = load_workbook('c:/Users/zhu_j/Desktop/BAOJIA/模板美林最新报价20260505.xlsx')
ws = wb.active

print('=== 工作表信息 ===')
print(f'最大行: {ws.max_row}, 最大列: {ws.max_column}')
print()
print('=== 前30行内容 ===')
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i > 30:
        break
    row_values = []
    for cell in row:
        if cell is not None:
            row_values.append(str(cell)[:40])
        else:
            row_values.append('')
    print(f'行{i}: {row_values}')
